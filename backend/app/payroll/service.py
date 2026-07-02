from sqlalchemy.orm import Session
from sqlalchemy import select, and_
from typing import Optional, List
from datetime import date, datetime, timezone
from decimal import Decimal
import calendar
import uuid

from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from app.payroll import repository
from app.payroll.model import SalaryHistory, SalaryComponent
from app.payroll.schema import PayrollCreate, PayrollUpdate, PayrollResponse, PayrollListResponse
from app.employees.model import Employee
from app.salary.repository import get_effective_for_month
from app.attendance.model import Attendance

from app.core.config import settings

def _enrich(db: Session, record: SalaryHistory) -> PayrollResponse:
    emp = db.execute(select(Employee).where(Employee.id == record.employee_id)).scalar_one_or_none()
    data = PayrollResponse.model_validate(record)
    if emp:
        data.employee_name = emp.name
        data.designation = emp.designation or ""
    return data


def list_payroll(
    db: Session,
    month: Optional[int],
    year: Optional[int],
    status_filter: str,
    employee_ids: Optional[List[str]],
    historical: bool,
    sort_by: str,
    sort_dir: str,
    page: int,
    page_size: int,
) -> PayrollListResponse:
    skip = (page - 1) * page_size
    total, items = repository.get_list(
        db, month, year, status_filter, employee_ids,
        historical, sort_by, sort_dir, skip, page_size
    )
    return PayrollListResponse(total=total, items=[_enrich(db, r) for r in items])

def list_my_payroll(
    db: Session,
    employee_id: str,
    month: Optional[int],
    year: Optional[int],
    sort_by: str,
    sort_dir: str,
    page: int,
    page_size: int,
) -> PayrollListResponse:
    skip = (page - 1) * page_size
    total, items = repository.get_list(
        db, month, year, "all", [employee_id],
        True, sort_by, sort_dir, skip, page_size
    )
    return PayrollListResponse(total=total, items=[_enrich(db, r) for r in items])



def create_payroll(db: Session, payload: PayrollCreate) -> PayrollResponse:
    emp = db.execute(select(Employee).where(Employee.id == payload.employee_id)).scalar_one_or_none()
    if not emp:
        raise ValueError("Employee not found")
    if repository.exists_for_employee_month(db, payload.employee_id, payload.month, payload.year):
        raise ValueError("Payroll record already exists for this employee and month")
    data = payload.model_dump()
    data["status"] = "pending"
    record = repository.create(db, data)
    db.commit()
    db.refresh(record)
    return _enrich(db, record)


def update_payroll(db: Session, record_id: uuid.UUID, payload: PayrollUpdate) -> PayrollResponse:
    record = repository.get_by_id(db, record_id)
    if not record:
        raise ValueError("Record not found")
    if record.status not in ("pending", "calculated"):
        raise ValueError("Cannot edit a paid record")
    data = payload.model_dump(exclude_unset=True)
    repository.update(db, record, data)
    db.commit()
    db.refresh(record)
    return _enrich(db, record)


def calculate_payroll(db: Session, record_id: uuid.UUID) -> PayrollResponse:
    record = repository.get_by_id(db, record_id)
    if not record:
        raise ValueError("Record not found")

    days_in_month = calendar.monthrange(record.year, record.month)[1]
    month_start = date(record.year, record.month, 1)
    month_end = date(record.year, record.month, days_in_month)

    struct = get_effective_for_month(db, record.employee_id, month_start, month_end)
    if not struct:
        raise ValueError("No effective salary structure found for this employee and month")

    # Attendance counts
    att_rows = db.execute(
        select(Attendance.status).where(
            and_(
                Attendance.employee_id == record.employee_id,
                Attendance.date >= month_start,
                Attendance.date <= month_end,
            )
        )
    ).scalars().all()

    days_present = sum(1 for s in att_rows if s in ("P", "PL", "SL", "Present", "Paid Leave", "Sick Leave"))
    days_absent = sum(1 for s in att_rows if s in ("A", "Absent"))
    leaves_taken = sum(1 for s in att_rows if s in ("PL", "SL", "Paid Leave", "Sick Leave"))

    def d(v):
        return Decimal(str(v)) if v is not None else Decimal("0")

    basic = d(struct.basic_allowance)
    hra = d(struct.hra_allowance)
    conv = d(struct.conveyance_allowance)
    med = d(struct.medical_allowance)
    dp = Decimal(str(days_present))
    dim = Decimal(str(days_in_month))
    ot_h = d(record.ot_hours)
    incentive = d(record.incentive)
    adv = d(record.advance)
    loan = d(record.loan)
    tds = d(record.tds)
    emp_mlwf = d(record.employee_mlwf)
    er_mlwf = d(record.employer_mlwf)

    ern_basic = (basic / dim * dp).quantize(Decimal("0.01"))
    ern_hra = (hra / dim * dp).quantize(Decimal("0.01"))
    ern_conv = (conv / dim * dp).quantize(Decimal("0.01"))
    ern_med = (med / dim * dp).quantize(Decimal("0.01"))

    if days_present > 0:
        ot_amount = (ern_basic * ot_h / (Decimal("8") * dp)).quantize(Decimal("0.01"))
    else:
        ot_amount = Decimal("0")

    gross = (ern_basic + ern_hra + ern_conv + ern_med + ot_amount + incentive).quantize(Decimal("0.01"))

    emp_pf = (Decimal("0.12") * ern_basic).quantize(Decimal("0.01"))
    emp_esic = Decimal("0") if ern_basic > 25000 else (Decimal("0.0075") * ern_basic).quantize(Decimal("0.01"))
    pt = Decimal("300") if record.month == 3 else Decimal("200")

    total_deductions = (emp_pf + emp_esic + pt + emp_mlwf + adv + loan + tds).quantize(Decimal("0.01"))
    net_salary = (gross - total_deductions).quantize(Decimal("0.01"))

    employer_pf = (Decimal("0.12") * ern_basic).quantize(Decimal("0.01"))
    employer_admin = Decimal("0") if ern_basic > 25000 else min((Decimal("0.01") * ern_basic).quantize(Decimal("0.01")), Decimal("1800"))
    employer_total_pf = (employer_pf + employer_admin).quantize(Decimal("0.01"))
    emp_employer_pf = (emp_pf + employer_total_pf).quantize(Decimal("0.01"))
    employer_esic = Decimal("0") if gross > 21000 else (Decimal("0.0325") * gross).quantize(Decimal("0.01"))
    emp_employer_esic = (emp_esic + employer_esic).quantize(Decimal("0.01"))
    emp_employer_mlwf = (emp_mlwf + er_mlwf).quantize(Decimal("0.01"))
    total_ctc = (net_salary + emp_employer_pf + emp_employer_esic + emp_employer_mlwf).quantize(Decimal("0.01"))

    updates = {
        "days_present": days_present,
        "days_absent": days_absent,
        "leaves_taken": leaves_taken,
        "ern_basic": ern_basic,
        "ern_hra": ern_hra,
        "ern_conveyance": ern_conv,
        "ern_medical": ern_med,
        "ot_amount": ot_amount,
        "gross_salary": gross,
        "emp_pf": emp_pf,
        "emp_esic": emp_esic,
        "pt": pt,
        "total_deductions": total_deductions,
        "net_salary": net_salary,
        "employer_pf": employer_pf,
        "employer_admin": employer_admin,
        "employer_total_pf": employer_total_pf,
        "emp_employer_pf": emp_employer_pf,
        "employer_esic": employer_esic,
        "emp_employer_esic": emp_employer_esic,
        "emp_employer_mlwf": emp_employer_mlwf,
        "total_ctc": total_ctc,
        "status": "calculated",
        "calculated_at": datetime.now(timezone.utc),
    }
    repository.update(db, record, updates)

    # Write salary components
    db.query(SalaryComponent).filter(SalaryComponent.salary_history_id == record.id).delete()
    components = [
        ("Basic", "earning", ern_basic),
        ("HRA", "earning", ern_hra),
        ("Conveyance", "earning", ern_conv),
        ("Medical", "earning", ern_med),
        ("OT Amount", "earning", ot_amount),
        ("Incentive", "earning", incentive),
        ("Employee PF", "deduction", emp_pf),
        ("Employee ESIC", "deduction", emp_esic),
        ("PT", "deduction", pt),
        ("Advance", "deduction", adv),
        ("Loan", "deduction", loan),
        ("TDS", "deduction", tds),
        ("Employee MLWF", "deduction", emp_mlwf),
    ]
    for name, ctype, amount in components:
        db.add(SalaryComponent(
            salary_history_id=record.id,
            component_name=name,
            component_type=ctype,
            amount=amount,
        ))

    db.commit()
    db.refresh(record)
    return _enrich(db, record)


def mark_paid(db: Session, record_id: uuid.UUID) -> PayrollResponse:
    record = repository.get_by_id(db, record_id)
    if not record:
        raise ValueError("Record not found")
    repository.update(db, record, {"status": "paid", "paid_at": datetime.now(timezone.utc)})
    db.commit()
    db.refresh(record)
    return _enrich(db, record)

VALID_TXN_TYPES = {"WIB", "NFT", "RTG", "IFC"}


def _validate_bank_row(net_salary, struct):
    errors = []
    if not struct:
        return ["No active salary structure for this month"]
    if struct.transaction_type not in VALID_TXN_TYPES:
        errors.append("Invalid or missing transaction type")
    if not struct.bene_id:
        errors.append("Missing Bene ID")
    if not struct.remarks or len(struct.remarks) > 30:
        errors.append("Remarks missing or exceeds 30 characters")
    if net_salary is None or net_salary <= 0:
        errors.append("Net salary missing or not positive")
    elif len(str(int(net_salary))) + 3 > 15:
        errors.append("Amount exceeds 15 digits")
    return errors


def _fmt_amount(amount: Decimal) -> str:
    amount = amount.quantize(Decimal("0.01"))
    return str(int(amount)) if amount == amount.to_integral_value() else str(amount)


def bank_export_data(db: Session, month: int, year: int):
    days_in_month = calendar.monthrange(year, month)[1]
    ref_date = date(year, month, days_in_month)
    rows = repository.get_calculated_for_month(db, month, year)

    included, skipped = [], []
    for record, emp_name in rows:
        struct = get_effective_for_month(db, record.employee_id, ref_date, ref_date)
        errors = _validate_bank_row(record.net_salary, struct)
        if errors:
            skipped.append({
                "employee_id": str(record.employee_id),
                "employee_name": emp_name,
                "reason": "; ".join(errors),
            })
            continue
        included.append({
            "employee_id": str(record.employee_id),
            "employee_name": emp_name,
            "transaction_type": struct.transaction_type,
            "amount": _fmt_amount(record.net_salary),
            "bene_id": struct.bene_id,
            "remarks": struct.remarks,
        })
    return included, skipped


def bank_export_file_text(included: list, month: int, year: int) -> str:

    today = date.today()
    if date(year, month, 1) < date(today.year, today.month, 1):
        raise ValueError("Bank export is only available for the current or future months")

    days_in_month = calendar.monthrange(year, month)[1]
    payment_date = date(year, month, days_in_month)
    total_amount = sum(Decimal(i["amount"]) for i in included)
    count = len(included)
    header = (
        f"FHR|0011|{settings.BANK_DEBIT_ACCOUNT}|INR"
        f"|{_fmt_amount(total_amount)}|{count}"
        f"|{payment_date.strftime('%m/%d/%Y')}|{settings.BANK_FILE_REMARKS}^"
    )
    lines = [header] + [
        f"PRB|{i['transaction_type']}|{i['amount']}|INR|{i['bene_id']}|"
        f"{settings.BANK_DEBIT_ACCOUNT}|0011|{i['remarks']}|N|PRBNBB^"
        for i in included
    ]
    return "\n".join(lines)

MONTH_NAMES_FULL = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
]

# 40 columns, A..AN — matches the muster roll template layout
MUSTER_ROLL_HEADERS = [
    "Sr.\nNo.", "Emp Code", "Employee Name", "Designation", "Date of Birth",
    "Aadhar No.", "PAN No.", "Date of Joining", "Days Present", "Fix Rate",
    "Basic", "HRA", "Conveyance", "Medical",
    "ERN Basic", "ERN HRA", "ERN Conveyance", "ERN Medical",
    "OT Hours", "OT Amt + Incentive", "Gross Salary",
    "Employee PF", "Employee ESIC", "PT", "Advance", "Loan", "TDS", "Employee MLWF",
    "Total Deductions", "Net Salary", "",
    "Employer PF", "Employer Admin", "Employer Total PF", "Emp+Employer PF",
    "Employer ESIC", "Emp+Employer ESIC", "Employer MLWF", "Emp+Employer MLWF",
    "Total CTC",
]


def build_muster_roll_workbook(
    db: Session,
    month: int,
    year: int,
    employee_ids: List[uuid.UUID],
) -> BytesIO:
    rows = repository.get_records_for_export(db, month, year, employee_ids)
    if not rows:
        raise ValueError(
            "No calculated or paid payroll records found for the selected "
            f"employees in {MONTH_NAMES_FULL[month - 1]} {year}."
        )

    days_in_month = calendar.monthrange(year, month)[1]

    wb = Workbook()
    ws = wb.active
    ws.title = "Muster Roll"

    bold = Font(bold=True)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Rows 1-9: static title / company / principal-employer block
    ws.cell(row=1, column=1, value="MUSTER ROLL").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"Company: {settings.MUSTER_ROLL_COMPANY_NAME}")
    ws.cell(row=3, column=1, value=f"Address: {settings.MUSTER_ROLL_COMPANY_ADDRESS}")
    ws.cell(row=4, column=1, value=f"Principal Employer: {settings.MUSTER_ROLL_PRINCIPAL_NAME}")
    ws.cell(row=5, column=1, value=f"Principal Employer Address: {settings.MUSTER_ROLL_PRINCIPAL_ADDRESS}")
    ws.cell(row=6, column=1, value=f"Month: {MONTH_NAMES_FULL[month - 1]} {year}")
    ws.cell(row=7, column=1, value=f"Days in Month: {days_in_month}")
    ws.cell(row=8, column=1, value="")
    ws.cell(row=9, column=1, value="")

    for rng in ("A1:AN1", "A2:AN2", "A3:AN3", "A4:AN4", "A5:AN5", "A6:AN6", "A7:AN7"):
        ws.merge_cells(rng)

    # Row 10: header labels
    for idx, label in enumerate(MUSTER_ROLL_HEADERS, start=1):
        cell = ws.cell(row=10, column=idx, value=label)
        cell.font = bold
        cell.alignment = header_align

    # Row 11: numeric reference row (kept per template)
    for col_idx in range(7, 35):
        ws.cell(row=11, column=col_idx, value=col_idx)

    # Row 12 is left blank as a spacer, per template.

    start_row = 13
    for sr, (record, emp) in enumerate(rows, start=1):
        r = start_row + sr - 1
        struct = get_effective_for_month(
            db, record.employee_id,
            date(year, month, 1), date(year, month, days_in_month),
        )
        basic = struct.basic_allowance if struct else None
        hra = struct.hra_allowance if struct else None
        conv = struct.conveyance_allowance if struct else None
        med = struct.medical_allowance if struct else None
        fix_rate = None
        if struct:
            fix_rate = sum(v for v in (basic, hra, conv, med) if v is not None)

        ot_plus_incentive = None
        if record.ot_amount is not None or record.incentive is not None:
            ot_plus_incentive = (record.ot_amount or Decimal("0")) + (record.incentive or Decimal("0"))

        values = [
            sr, emp.emp_code, emp.name, emp.designation, emp.date_of_birth,
            emp.aadhar_no, emp.pan_no, emp.join_date, record.days_present, fix_rate,
            basic, hra, conv, med,
            record.ern_basic, record.ern_hra, record.ern_conveyance, record.ern_medical,
            record.ot_hours, ot_plus_incentive, record.gross_salary,
            record.emp_pf, record.emp_esic, record.pt, record.advance, record.loan,
            record.tds, record.employee_mlwf,
            record.total_deductions, record.net_salary, None,
            record.employer_pf, record.employer_admin, record.employer_total_pf,
            record.emp_employer_pf, record.employer_esic, record.emp_employer_esic,
            record.employer_mlwf, record.emp_employer_mlwf, record.total_ctc,
        ]
        for col_idx, val in enumerate(values, start=1):
            ws.cell(row=r, column=col_idx, value=val)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf